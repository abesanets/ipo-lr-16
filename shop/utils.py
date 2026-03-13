import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO


def generate_receipt_excel(order):
    """Генерация чека заказа в формате Excel"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Чек #{order.id}"

    # Стили
    header_font = Font(name='Arial', size=16, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=11)
    bold_font = Font(name='Arial', size=11, bold=True)
    total_font = Font(name='Arial', size=14, bold=True, color='FF0000')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_font_white = Font(name='Arial', size=11, bold=True, color='FFFFFF')

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # Ширина столбцов
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18

    # === Заголовок ===
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = '🛒 Магазин электроники — ЧЕК'
    cell.font = header_font
    cell.alignment = center_align

    # === Информация о заказе ===
    row = 3
    info_data = [
        ('Номер заказа:', f'#{order.id}'),
        ('Дата:', order.created_at.strftime('%d.%m.%Y %H:%M')),
        ('Покупатель:', order.user.get_full_name() or order.user.username),
        ('Email:', order.email),
        ('Телефон:', order.phone),
        ('Адрес доставки:', order.address),
    ]

    for label, value in info_data:
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=str(value)).font = normal_font
        row += 1

    # === Таблица товаров ===
    row += 1
    table_headers = ['№', 'Товар', 'Цена', 'Кол-во', 'Стоимость']

    for col, header in enumerate(table_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    row += 1
    for idx, item in enumerate(order.items.all(), 1):
        ws.cell(row=row, column=1, value=idx).font = normal_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=item.product_name).font = normal_font
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=float(item.product_price)).font = normal_font
        ws.cell(row=row, column=3).number_format = '#,##0.00 ₽'
        ws.cell(row=row, column=3).alignment = right_align
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4, value=item.quantity).font = normal_font
        ws.cell(row=row, column=4).alignment = center_align
        ws.cell(row=row, column=4).border = thin_border

        ws.cell(row=row, column=5, value=float(item.item_cost())).font = normal_font
        ws.cell(row=row, column=5).number_format = '#,##0.00 ₽'
        ws.cell(row=row, column=5).alignment = right_align
        ws.cell(row=row, column=5).border = thin_border

        row += 1

    # === Итого ===
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value='ИТОГО:').font = total_font
    ws.cell(row=row, column=1).alignment = right_align
    ws.cell(row=row, column=5, value=float(order.total_cost)).font = total_font
    ws.cell(row=row, column=5).number_format = '#,##0.00 ₽'
    ws.cell(row=row, column=5).alignment = right_align

    # === Футер ===
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value='Спасибо за покупку!').font = subheader_font
    ws.cell(row=row, column=1).alignment = center_align

    # Сохраняем в BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer